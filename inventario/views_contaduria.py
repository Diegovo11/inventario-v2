# ================ VISTAS PARA SISTEMA DE CONTADURÍA ================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from .models import MovimientoEfectivo
from .forms import MovimientoEfectivoForm, FiltroMovimientosEfectivoForm
from datetime import datetime, timedelta


@login_required
def contaduria_home(request):
    """Vista principal del sistema de contaduría"""
    # Obtener saldo actual
    saldo_actual = MovimientoEfectivo.calcular_saldo_actual()
    
    # Obtener últimos movimientos
    ultimos_movimientos = MovimientoEfectivo.objects.all()[:10]
    
    # Calcular estadísticas del mes actual
    inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    movimientos_mes = MovimientoEfectivo.objects.filter(
        fecha__gte=inicio_mes,
        fecha__lte=fin_mes
    )
    
    ingresos_mes = sum(m.monto for m in movimientos_mes if m.tipo_movimiento == 'ingreso')
    egresos_mes = sum(m.monto for m in movimientos_mes if m.tipo_movimiento == 'egreso')
    balance_mes = ingresos_mes - egresos_mes
    
    context = {
        'saldo_actual': saldo_actual,
        'ultimos_movimientos': ultimos_movimientos,
        'ingresos_mes': ingresos_mes,
        'egresos_mes': egresos_mes,
        'balance_mes': balance_mes,
        'title': 'Contaduría y Finanzas'
    }
    return render(request, 'inventario/contaduria_home.html', context)


@login_required
def flujo_efectivo(request):
    """Vista para mostrar el flujo de efectivo con filtros"""
    filtro_form = FiltroMovimientosEfectivoForm(request.GET or None)
    movimientos = MovimientoEfectivo.objects.all()
    
    if filtro_form.is_valid():
        if filtro_form.cleaned_data.get('fecha_inicio'):
            movimientos = movimientos.filter(fecha__date__gte=filtro_form.cleaned_data['fecha_inicio'])
        
        if filtro_form.cleaned_data.get('fecha_fin'):
            movimientos = movimientos.filter(fecha__date__lte=filtro_form.cleaned_data['fecha_fin'])
        
        if filtro_form.cleaned_data.get('tipo_movimiento'):
            movimientos = movimientos.filter(tipo_movimiento=filtro_form.cleaned_data['tipo_movimiento'])
        
        if filtro_form.cleaned_data.get('categoria'):
            movimientos = movimientos.filter(categoria=filtro_form.cleaned_data['categoria'])
        
        if filtro_form.cleaned_data.get('automatico'):
            es_automatico = filtro_form.cleaned_data['automatico'] == 'true'
            movimientos = movimientos.filter(automatico=es_automatico)
    
    # Paginación
    paginator = Paginator(movimientos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calcular totales
    total_ingresos = sum(m.monto for m in movimientos if m.tipo_movimiento == 'ingreso')
    total_egresos = sum(m.monto for m in movimientos if m.tipo_movimiento == 'egreso')
    balance = total_ingresos - total_egresos
    
    context = {
        'movimientos': page_obj,
        'filtro_form': filtro_form,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'balance_neto': balance,
        'title': 'Flujo de Efectivo'
    }
    return render(request, 'inventario/flujo_efectivo.html', context)


@login_required
def registrar_movimiento_efectivo(request):
    """Vista para registrar movimientos de efectivo manuales"""
    if request.method == 'POST':
        form = MovimientoEfectivoForm(request.POST)
        if form.is_valid():
            # Usar el método de clase para registrar el movimiento
            movimiento = MovimientoEfectivo.registrar_movimiento(
                concepto=form.cleaned_data['concepto'],
                tipo_movimiento=form.cleaned_data['tipo_movimiento'],
                categoria=form.cleaned_data['categoria'],
                monto=form.cleaned_data['monto'],
                usuario=request.user
            )
            
            tipo_texto = 'Ingreso' if movimiento.tipo_movimiento == 'ingreso' else 'Egreso'
            messages.success(
                request,
                f'{tipo_texto} registrado exitosamente. '
                f'Saldo actual: ${movimiento.saldo_nuevo:.2f}'
            )
            return redirect('inventario:flujo_efectivo')
    else:
        form = MovimientoEfectivoForm()
    
    context = {
        'form': form,
        'title': 'Registrar Movimiento de Efectivo'
    }
    return render(request, 'inventario/registrar_movimiento_efectivo.html', context)


@login_required
def estado_resultados(request):
    """Vista para generar estado de resultados"""
    # Obtener fechas del formulario o usar mes actual por defecto
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not fecha_inicio:
        inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        fecha_inicio = inicio_mes.date()
    else:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    
    if not fecha_fin:
        fin_mes = (datetime.now().replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        fecha_fin = fin_mes.date()
    else:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # Obtener movimientos del período
    movimientos = MovimientoEfectivo.objects.filter(
        fecha__date__gte=fecha_inicio,
        fecha__date__lte=fecha_fin
    )
    
    # Calcular totales generales
    total_ingresos = sum(m.monto for m in movimientos if m.tipo_movimiento == 'ingreso')
    total_egresos = sum(m.monto for m in movimientos if m.tipo_movimiento == 'egreso')
    resultado_neto = total_ingresos - total_egresos
    
    # Calcular ingresos por categoría
    ingresos_por_categoria = {}
    for mov in movimientos.filter(tipo_movimiento='ingreso'):
        categoria = mov.get_categoria_display()
        if categoria not in ingresos_por_categoria:
            ingresos_por_categoria[categoria] = {'total': 0, 'cantidad': 0}
        ingresos_por_categoria[categoria]['total'] += mov.monto
        ingresos_por_categoria[categoria]['cantidad'] += 1
    
    # Calcular egresos por categoría  
    egresos_por_categoria = {}
    for mov in movimientos.filter(tipo_movimiento='egreso'):
        categoria = mov.get_categoria_display()
        if categoria not in egresos_por_categoria:
            egresos_por_categoria[categoria] = {'total': 0, 'cantidad': 0}
        egresos_por_categoria[categoria]['total'] += mov.monto
        egresos_por_categoria[categoria]['cantidad'] += 1
    
    # Calcular análisis por origen
    ingresos_automaticos = sum(m.monto for m in movimientos if m.tipo_movimiento == 'ingreso' and m.automatico)
    egresos_automaticos = sum(m.monto for m in movimientos if m.tipo_movimiento == 'egreso' and m.automatico)
    ingresos_manuales = sum(m.monto for m in movimientos if m.tipo_movimiento == 'ingreso' and not m.automatico)
    egresos_manuales = sum(m.monto for m in movimientos if m.tipo_movimiento == 'egreso' and not m.automatico)
    neto_automaticos = ingresos_automaticos - egresos_automaticos
    neto_manuales = ingresos_manuales - egresos_manuales
    
    # Estadísticas adicionales
    total_movimientos = movimientos.count()
    promedio_ingresos = total_ingresos / max(movimientos.filter(tipo_movimiento='ingreso').count(), 1)
    promedio_egresos = total_egresos / max(movimientos.filter(tipo_movimiento='egreso').count(), 1)
    margen_neto = (resultado_neto / max(total_ingresos, 1)) * 100 if total_ingresos > 0 else 0
    
    # Crear formulario de filtros
    filtro_form = FiltroMovimientosEfectivoForm(initial={
        'fecha_desde': fecha_inicio,
        'fecha_hasta': fecha_fin
    })
    
    context = {
        'fecha_desde': fecha_inicio,
        'fecha_hasta': fecha_fin,
        'filtro_form': filtro_form,
        'ingresos_por_categoria': ingresos_por_categoria,
        'egresos_por_categoria': egresos_por_categoria,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'resultado_neto': resultado_neto,
        'ingresos_automaticos': ingresos_automaticos,
        'egresos_automaticos': egresos_automaticos,
        'ingresos_manuales': ingresos_manuales,
        'egresos_manuales': egresos_manuales,
        'neto_automaticos': neto_automaticos,
        'neto_manuales': neto_manuales,
        'total_movimientos': total_movimientos,
        'promedio_ingresos': promedio_ingresos,
        'promedio_egresos': promedio_egresos,
        'margen_neto': margen_neto,
        'title': 'Estado de Resultados'
    }
    return render(request, 'inventario/estado_resultados.html', context)


@login_required 
def exportar_excel_efectivo(request):
    """Exportar movimientos de efectivo a Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    
    # Crear libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flujo de Efectivo"
    
    # Encabezados
    headers = ['Fecha', 'Concepto', 'Tipo', 'Categoría', 'Monto', 'Saldo', 'Origen']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    movimientos = MovimientoEfectivo.objects.all()[:1000]  # Limitar a 1000 registros
    
    for row, movimiento in enumerate(movimientos, 2):
        ws.cell(row=row, column=1, value=movimiento.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=movimiento.concepto)
        ws.cell(row=row, column=3, value=movimiento.get_tipo_movimiento_display())
        ws.cell(row=row, column=4, value=movimiento.get_categoria_display())
        
        # Monto con signo
        monto_signo = movimiento.monto if movimiento.tipo_movimiento == 'ingreso' else -movimiento.monto
        ws.cell(row=row, column=5, value=float(monto_signo))
        ws.cell(row=row, column=6, value=float(movimiento.saldo_nuevo))
        ws.cell(row=row, column=7, value='Automático' if movimiento.automatico else 'Manual')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta HTTP
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="flujo_efectivo_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    return response